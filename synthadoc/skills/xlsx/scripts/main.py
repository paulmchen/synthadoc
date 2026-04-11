# SPDX-License-Identifier: AGPL-3.0-or-later
# Copyright (C) 2026 Paul Chen / axoviq.com
import openpyxl
from synthadoc.skills.base import BaseSkill, ExtractedContent, SkillMeta


class XlsxSkill(BaseSkill):
    meta = SkillMeta(name="xlsx", description="Extract data from Excel and CSV files",
                     extensions=[".xlsx", ".csv"])

    async def extract(self, source: str) -> ExtractedContent:
        if source.endswith(".csv"):
            import csv
            rows = []
            with open(source, encoding="utf-8", newline="") as f:
                for row in csv.reader(f):
                    rows.append(", ".join(row))
            return ExtractedContent(text="\n".join(rows), source_path=source, metadata={})
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
        parts = []
        for name in wb.sheetnames:
            parts.append(f"## Sheet: {name}")
            for row in wb[name].iter_rows(values_only=True):
                parts.append(", ".join(str(c) for c in row if c is not None))
        return ExtractedContent(text="\n".join(parts), source_path=source,
                                metadata={"sheets": len(wb.sheetnames)})
